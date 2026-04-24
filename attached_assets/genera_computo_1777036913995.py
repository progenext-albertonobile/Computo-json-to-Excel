#!/usr/bin/env python3
"""
genera_computo.py — Progenext | Generatore computo metrico estimativo
Versione 2.1 — architettura data-driven con type per riga
Data:    2026-04-24

Usage:  python3 genera_computo.py input.json [output.xlsx]

Tipi riga supportati:
  sezione           — intestazione sezione (navy)
  sottosezione      — intestazione sottogruppo (azzurro chiaro, bg opzionale)
  voce              — riga standard con formule complete (bianco/grigio alt)
  voce_highlight    — voce evidenziata (verde)
  voce_demolizione  — voce demolizione (rosa)
  voce_trasporto    — trasporto semplificato (bianco/grigio alt)
  voce_lumpsum      — prezzo a corpo, col R hardcoded (bianco/grigio alt)
  voce_esterna      — fuori totale (arancio)
  sottotot_sezione  — subtotale sezione automatico (blu)
  totale            — totale generale lavori (navy)
  esclusioni        — intestazione sezione esclusioni (navy)
  riga_vuota        — riga vuota di separazione (altezza 6pt)
"""
import json, sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# ── PALETTE ──────────────────────────────────────────────────────────────────
C = {
    'navy':   'FF1A3A5C',
    'blue':   'FF2563AE',
    'lbhdr':  'FF4A90D9',
    'lbbg':   'FFDCE9F8',
    'green':  'FFD6EED9',
    'pink':   'FFFADBD8',
    'orange': 'FFFAE3D0',
    'gray':   'FFF5F5F5',
    'white':  'FFFFFFFF',
    'black':  'FF000000',
    'fwhite': 'FFFFFFFF',
    'fnavy':  'FF1A3A5C',
    'fgray':  'FF606060',
}

# ── COLUMN LAYOUT ─────────────────────────────────────────────────────────────
# 19 columns: A=desc B=brand C=specs D=qty E=pu F=fornitura G=sc% H=mat
# I=op% J=op€ K=inst% L=inst€ M=ric% N=ric€ O=tot_riga P=impr% Q=impr€ R=prezzo S=note
NCOLS = 19
COL_W = {
    1: 36,   # A  desc
    2: 16,   # B  brand
    3: 48,   # C  specs
    4: 5,    # D  qty
    5: 11,   # E  P.U.
    6: 11,   # F  fornitura
    7: 5,    # G  sc%
    8: 11,   # H  materiale
    9: 5,    # I  op.acc%
    10: 9,   # J  op.acc€
    11: 5,   # K  inst%
    12: 9,   # L  inst€
    13: 5,   # M  ric%
    14: 9,   # N  ric€
    15: 11,  # O  TOT riga
    16: 5,   # P  impr%
    17: 9,   # Q  impr€
    18: 13,  # R  PREZZO TOTALE
    19: 36,  # S  note
}
EUR = '#,##0;(#,##0);"-"'
PCT = '0%'
PCT_COLS  = {7, 9, 11, 13, 16}
EUR_COLS  = {4, 5, 6, 8, 10, 12, 14, 15, 17, 18}
TEXT_COLS = {1, 2, 3, 19}

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
def _fill(rgb):
    return PatternFill('solid', fgColor=rgb)

def _font(rgb, bold=True, sz=8):
    return Font(color=rgb, bold=bold, size=sz, name='Calibri')

def _align(h='left', wrap=True):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

def style(cell, bg, fc, bold=True, sz=8, h='left', wrap=True, nfmt=None):
    cell.fill      = _fill(bg)
    cell.font      = _font(fc, bold, sz)
    cell.alignment = _align(h, wrap)
    if nfmt:
        cell.number_format = nfmt

def fill_row(ws, row, bg, fc, bold=True, sz=8):
    for c in range(1, NCOLS + 1):
        style(ws.cell(row, c), bg, fc, bold, sz)

def set_val(ws, row, col, val, bg, fc, bold=True, sz=8, h=None, wrap=True):
    is_text = col in TEXT_COLS
    auto_h  = 'left' if is_text else 'right'
    cell = ws.cell(row, col, val)
    style(cell, bg, fc, bold, sz, h or auto_h, wrap)
    if col in EUR_COLS:
        cell.number_format = EUR
    elif col in PCT_COLS:
        cell.number_format = PCT
    return cell

# ── ROW RENDERERS ─────────────────────────────────────────────────────────────

def render_sezione(ws, r, row_data, label=None):
    ws.row_dimensions[r].height = 16
    fill_row(ws, r, C['navy'], C['fwhite'], sz=9)
    ws.merge_cells(f'A{r}:S{r}')
    txt = label or row_data.get('label', '')
    style(ws.cell(r, 1, f'  ━━━━━━  {txt}  ━━━━━━'),
          C['navy'], C['fwhite'], sz=9)

def render_sottosezione(ws, r, row_data):
    ws.row_dimensions[r].height = 13
    bg = row_data.get('bg') or C['lbbg']
    fill_row(ws, r, bg, C['fnavy'], sz=8)
    ws.merge_cells(f'A{r}:S{r}')
    style(ws.cell(r, 1, f'  {row_data.get("label","")}'),
          bg, C['fnavy'], sz=8)

def render_voce_base(ws, r, row_data, bg, alt_tracker):
    """Core renderer for voce, voce_highlight, voce_demolizione, voce_esterna."""
    ws.row_dimensions[r].height = 38
    rn = r
    sc   = row_data.get('sc', 0)
    oa   = row_data.get('op_acc', 0)
    inst = row_data.get('inst', 0)
    ric  = row_data.get('ric', 0)
    impr = row_data.get('impr', 0)

    vals = {
        1:  row_data.get('desc', ''),
        2:  row_data.get('brand', ''),
        3:  row_data.get('specs', ''),
        4:  row_data.get('qty', 1),
        5:  row_data.get('pu', 0),
        6:  f'=D{rn}*E{rn}',
        7:  sc,
        8:  f'=F{rn}*(1-G{rn})',
        9:  oa,
        10: f'=H{rn}*I{rn}',
        11: inst,
        12: f'=(H{rn}+J{rn})*K{rn}',
        13: ric,
        14: f'=(H{rn}+J{rn}+L{rn})*M{rn}',
        15: f'=H{rn}+J{rn}+L{rn}+N{rn}',
        16: impr,
        17: f'=O{rn}*P{rn}',
        18: f'=O{rn}+Q{rn}',
        19: row_data.get('note', ''),
    }
    for col, val in vals.items():
        set_val(ws, rn, col, val, bg, C['black'])

def render_voce(ws, r, row_data, alt):
    bg = C['white'] if alt % 2 == 0 else C['gray']
    render_voce_base(ws, r, row_data, bg, alt)

def render_voce_highlight(ws, r, row_data, _alt):
    render_voce_base(ws, r, row_data, C['green'], _alt)

def render_voce_demolizione(ws, r, row_data, _alt):
    render_voce_base(ws, r, row_data, C['pink'], _alt)

def render_voce_esterna(ws, r, row_data, _alt):
    render_voce_base(ws, r, row_data, C['orange'], _alt)

def render_voce_trasporto(ws, r, row_data, alt):
    ws.row_dimensions[r].height = 20
    bg  = C['white'] if alt % 2 == 0 else C['gray']
    rn  = r
    ric  = row_data.get('ric', 0.12)
    impr = row_data.get('impr', 0)

    vals = {
        1:  row_data.get('desc', ''),
        2:  row_data.get('brand', ''),
        3:  row_data.get('specs', ''),
        4:  row_data.get('qty', 1),
        5:  row_data.get('pu', 0),
        6:  f'=D{rn}*E{rn}',
        7:  0,
        8:  f'=F{rn}*(1-G{rn})',
        9:  0, 10: 0, 11: 0, 12: 0,
        13: ric,
        14: f'=H{rn}*M{rn}',
        15: f'=H{rn}+N{rn}',
        16: impr,
        17: f'=O{rn}*P{rn}',
        18: f'=O{rn}+Q{rn}',
        19: row_data.get('note', ''),
    }
    for col, val in vals.items():
        set_val(ws, rn, col, val, bg, C['black'])

def render_voce_lumpsum(ws, r, row_data, alt):
    ws.row_dimensions[r].height = 28
    bg  = C['white'] if alt % 2 == 0 else C['gray']
    rn  = r
    total = row_data.get('total', 0)

    vals = {
        1:  row_data.get('desc', ''),
        2:  row_data.get('brand', ''),
        3:  row_data.get('specs', ''),
        4:  row_data.get('qty', 1),
        5:  0, 6: 0, 7: 0, 8: 0,
        9:  0, 10: 0, 11: 0, 12: 0,
        13: 0, 14: 0, 15: 0, 16: 0, 17: 0,
        18: total,    # hardcoded lump-sum in R
        19: row_data.get('note', ''),
    }
    for col, val in vals.items():
        set_val(ws, rn, col, val, bg, C['black'])

def render_esclusioni(ws, r, row_data):
    ws.row_dimensions[r].height = 14
    fill_row(ws, r, C['navy'], C['fwhite'], sz=9)
    ws.merge_cells(f'A{r}:S{r}')
    style(ws.cell(r, 1, f'  {row_data.get("label","")}'),
          C['navy'], C['fwhite'], sz=9)

def render_riga_vuota(ws, r):
    ws.row_dimensions[r].height = 6
    fill_row(ws, r, C['white'], C['white'])

def render_subtot(ws, r, label, r_start, r_end):
    ws.row_dimensions[r].height = 18
    fill_row(ws, r, C['blue'], C['fwhite'], sz=9)
    ws.merge_cells(f'A{r}:Q{r}')
    style(ws.cell(r, 1, f'  SUBTOTALE {label} — IVA esclusa'),
          C['blue'], C['fwhite'], sz=9)
    c = ws.cell(r, 18, f'=SUM(R{r_start}:R{r_end})')
    style(c, C['blue'], C['fwhite'], sz=11, h='right')
    c.number_format = EUR
    # S: clear
    style(ws.cell(r, 19, ''), C['blue'], C['fwhite'])

def render_totale(ws, r, subtot_rows):
    ws.row_dimensions[r].height = 20
    fill_row(ws, r, C['navy'], C['fwhite'], sz=10)
    ws.merge_cells(f'A{r}:Q{r}')
    style(ws.cell(r, 1, '  TOTALE LAVORI — IVA esclusa'),
          C['navy'], C['fwhite'], sz=10)
    formula = '=SUM(' + ','.join(f'R{sr}' for sr in subtot_rows) + ')'
    c = ws.cell(r, 18, formula)
    style(c, C['navy'], C['fwhite'], sz=12, h='right')
    c.number_format = EUR
    style(ws.cell(r, 19, ''), C['navy'], C['fwhite'])

# ── FIXED HEADER (rows 1-4) ──────────────────────────────────────────────────

def build_header(ws, meta):
    # Row 1 — title + date
    ws.row_dimensions[1].height = 22
    fill_row(ws, 1, C['navy'], C['fwhite'], sz=10)
    ws.merge_cells('A1:R1')
    style(ws.cell(1, 1,
          f'  STIMA PRELIMINARE — {meta["cliente"]}  |  {meta["indirizzo"]}'),
          C['navy'], C['fwhite'], sz=10)
    style(ws.cell(1, 19, meta.get('data', '')),
          C['navy'], C['fwhite'], sz=9, h='right')

    # Row 2 — scenario title + note header
    ws.row_dimensions[2].height = 20
    fill_row(ws, 2, C['blue'], C['fwhite'], sz=10)
    ws.merge_cells('A2:L2')
    style(ws.cell(2, 1, f'  {meta["titolo"]}'),
          C['blue'], C['fwhite'], sz=10)
    ws.merge_cells('M2:S2')
    style(ws.cell(2, 13, meta.get('note_header', '')),
          C['blue'], C['fwhite'], sz=7, h='center')

    # Row 3 — column headers
    ws.row_dimensions[3].height = 30
    hdrs = [
        'Articolo / Descrizione', 'Marca / Rif.', 'Specifiche tecniche e note',
        'q.ta', 'P.U. netto\ninstall. €', 'Importo\nfornitura €', 'sc.%',
        'Costo\nmateriale €', 'op.\nacc%', 'op.acc\n€',
        'inst.\n%', 'install\n€', 'ric.\n%', 'ricarico\n€',
        'TOT\nriga €', 'impr\n%', 'impr.\n€',
        'PREZZO\nTOTALE €', 'Nota / Fonte / Affidabilità',
    ]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(3, i, h)
        style(c, C['lbhdr'], C['fwhite'], sz=7.5, h='center')

    # Row 4 — coefficient note
    ws.row_dimensions[4].height = 11
    ws.merge_cells('A4:H4')
    style(ws.cell(4, 1,
          '  Coefficienti individuali per riga — vedi colonne I, K, M, P'),
          C['gray'], C['fgray'], bold=False, sz=6.5)
    for col, txt in [(9,'↓ riga'),(11,'↓ riga'),(13,'↓ riga'),(16,'↓ riga')]:
        style(ws.cell(4, col, txt), C['gray'], C['fgray'],
              bold=False, sz=6.5, h='center')
    # Style remaining cells in row 4 (skip merged A4:H4)
    for col in [10,12,14,15,17,18,19]:
        c = ws.cell(4, col)
        c.fill = _fill(C['gray'])
        c.font = _font(C['fgray'], bold=False, sz=6.5)

# ── MAIN BUILD ────────────────────────────────────────────────────────────────

def build(data):
    meta = data.get('metadata', {})
    rows = data.get('rows', [])

    wb = Workbook()
    ws = wb.active
    ws.title = 'COMPUTO'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A5'

    for col, w in COL_W.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    build_header(ws, meta)

    cur = 5           # current Excel row (1-indexed), start after header
    alt = 0           # alternating counter for bianco/grigio
    subtot_rows = []  # Excel row numbers of sottotot_sezione rows
    sec_start = 5     # first data row of current section (for SUM range)
    totale_row = None

    for row_data in rows:
        t = row_data.get('type', '')

        if t == 'sezione':
            render_sezione(ws, cur, row_data)
            sec_start = cur + 1
            alt = 0

        elif t == 'sottosezione':
            render_sottosezione(ws, cur, row_data)

        elif t == 'voce':
            render_voce(ws, cur, row_data, alt); alt += 1

        elif t == 'voce_highlight':
            render_voce_highlight(ws, cur, row_data, alt); alt += 1

        elif t == 'voce_demolizione':
            render_voce_demolizione(ws, cur, row_data, alt); alt += 1

        elif t == 'voce_esterna':
            render_voce_esterna(ws, cur, row_data, alt); alt += 1

        elif t == 'voce_trasporto':
            render_voce_trasporto(ws, cur, row_data, alt); alt += 1

        elif t == 'voce_lumpsum':
            render_voce_lumpsum(ws, cur, row_data, alt); alt += 1

        elif t == 'sottotot_sezione':
            render_subtot(ws, cur, row_data.get('label', ''), sec_start, cur - 1)
            subtot_rows.append(cur)
            sec_start = cur + 1   # reset for next section

        elif t == 'totale':
            render_totale(ws, cur, subtot_rows)
            totale_row = cur

        elif t == 'esclusioni':
            render_esclusioni(ws, cur, row_data)

        elif t == 'riga_vuota':
            render_riga_vuota(ws, cur)

        else:
            # Unknown type — render as plain gray row with warning
            ws.row_dimensions[cur].height = 12
            fill_row(ws, cur, C['gray'], C['fgray'])
            ws.cell(cur, 1, f'⚠️ tipo sconosciuto: {t}')

        cur += 1

    # ── Nascondi righe vuote sotto il contenuto ──────────────────────────────
    # openpyxl non tocca le righe non usate: Excel le mostra come spazio bianco.
    # Impostiamo height=0 (hidden) per tutte le righe oltre l'ultimo contenuto.
    for r in range(cur, cur + 300):
        ws.row_dimensions[r].hidden = True

    # ── Named ranges for cross-sheet reference ──
    for i, sr in enumerate(subtot_rows, 1):
        nb = DefinedName(f'SUBTOT_S{i}', attr_text=f"'COMPUTO'!$R${sr}")
        wb.defined_names[f'SUBTOT_S{i}'] = nb
    if totale_row:
        wb.defined_names['TOTALE_COMPUTO'] = DefinedName(
            'TOTALE_COMPUTO', attr_text=f"'COMPUTO'!$R${totale_row}")

    return wb

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 genera_computo.py input.json [output.xlsx]")
        sys.exit(1)
    with open(sys.argv[1], encoding='utf-8') as f:
        data = json.load(f)
    wb = build(data)
    out = sys.argv[2] if len(sys.argv) > 2 else 'COMPUTO_OUTPUT.xlsx'
    wb.save(out)
    print(f"✅ Salvato: {out}")
