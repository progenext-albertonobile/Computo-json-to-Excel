#!/usr/bin/env python3
"""
genera_computo.py
Genera l'xlsx partendo dal template reale "Computo preliminare-V3.xlsx".
Usage:
  python genera_computo.py input.json [output.xlsx] [template.xlsx]
  python genera_computo.py --stamp-template-version v3.0.0 [--template template.xlsx]
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

SCENARIO_SHEET = "SC_C_FULL"
RIEPILOGO_SHEET = "RIEPILOGO"
_LOCAL_TEMPLATE = Path(__file__).resolve().parent / "templates" / "Computo preliminare-V3.xlsx"
_LEGACY_TEMPLATE = Path(__file__).resolve().parent.parent / "Computo preliminare-V3.xlsx"
DEFAULT_TEMPLATE = _LOCAL_TEMPLATE if _LOCAL_TEMPLATE.exists() else _LEGACY_TEMPLATE
TEMPLATE_VERSION_NAME = "COMPUTO_TEMPLATE_VERSION"
OUTPUT_GENERATOR_NAME = "COMPUTO_GENERATOR_VERSION"
GENERATOR_VERSION = "genera_computo.py@2026-04-16"

INPUT_COLS_TEXT = ("A", "B", "C", "S")
INPUT_COLS_NUM = ("D", "E", "G", "I", "K", "M", "P")


def _norm(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _norm_loose(value: object) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def _contains(haystack: object, needle: object) -> bool:
    strict = _norm(needle) in _norm(haystack)
    if strict:
        return True
    return _norm_loose(needle) in _norm_loose(haystack)


def _find_row_by_a_contains(ws, text: str, start: int = 1, end: int | None = None) -> int | None:
    last = end or ws.max_row
    for row in range(start, last + 1):
        if _contains(ws[f"A{row}"].value, text):
            return row
    return None


def _find_header_row(ws) -> int:
    row = _find_row_by_a_contains(ws, "Articolo / Descrizione")
    if not row:
        raise ValueError("Header 'Articolo / Descrizione' non trovato nel template")
    return row


def _find_section_rows(ws, sections: list[dict]) -> list[int]:
    rows: list[int] = []
    cursor = 1
    for sec_idx, sec in enumerate(sections):
        found = _find_row_by_a_contains(ws, sec.get("label", ""), start=cursor)
        if not found:
            for row in range(cursor, ws.max_row + 1):
                txt = str(ws[f"A{row}"].value or "")
                if ("SEZIONE" in txt) or ("PRESTAZIONI TECNICHE" in txt):
                    found = row
                    break
        if not found:
            raise ValueError(f"Sezione {sec_idx + 1} non trovata nel template: {sec.get('label', '')}")
        rows.append(found)
        cursor = found + 1
    return rows


def _find_subsection_rows(ws, section_row: int, section_end: int, subsections: list[dict]) -> list[int]:
    rows: list[int] = []
    cursor = section_row + 1
    for sub in subsections:
        found = _find_row_by_a_contains(ws, sub.get("label", ""), start=cursor, end=section_end)
        if not found:
            for row in range(cursor, section_end):
                a_val = str(ws[f"A{row}"].value or "")
                next_d = ws[f"D{row + 1}"].value if (row + 1) <= section_end else None
                if (
                    True
                    and ("SUBTOTALE" not in a_val)
                    and ("SEZIONE" not in a_val)
                    and ws[f"D{row}"].value is None
                    and next_d is not None
                ):
                    found = row
                    break
        if not found:
            raise ValueError(f"Sottosezione non trovata nel template: {sub.get('label', '')}")
        rows.append(found)
        cursor = found + 1
    return rows


def _find_section_subtotal_row(ws, section_row: int, next_section_row: int | None = None) -> int:
    end = (next_section_row - 1) if next_section_row else ws.max_row
    for row in range(section_row + 1, end + 1):
        if _contains(ws[f"A{row}"].value, "SUBTOTALE"):
            return row
    raise ValueError(f"Subtotale non trovato per sezione a riga {section_row}")


def _find_item_rows_in_section(ws, section_row: int, section_end: int, prefer_lumpsum: bool) -> list[int]:
    rows: list[int] = []
    for row in range(section_row + 1, section_end):
        d_val = ws[f"D{row}"].value
        a_val = ws[f"A{row}"].value
        if d_val is None or not str(a_val or "").strip():
            continue
        if str(a_val or "").startswith("  "):
            continue
        if prefer_lumpsum:
            r_val = ws[f"R{row}"].value
            if isinstance(r_val, (int, float)):
                rows.append(row)
        else:
            rows.append(row)
    return rows


def _quoted(value: str) -> str:
    return f'"{value}"'


def _unquote(value: str | None) -> str | None:
    if value is None:
        return None
    if len(value) >= 2 and value[0] == '"' and value[-1] == '"':
        return value[1:-1]
    return value


def _get_defined_name_value(wb, name: str) -> str | None:
    if name not in wb.defined_names:
        return None
    return _unquote(wb.defined_names[name].attr_text)


def _set_defined_name_value(wb, name: str, value: str) -> None:
    wb.defined_names[name] = DefinedName(name=name, attr_text=_quoted(value))


def stamp_template_version(template_path: Path, version: str) -> None:
    wb = load_workbook(template_path)
    _set_defined_name_value(wb, TEMPLATE_VERSION_NAME, version)
    wb.save(template_path)


def _flatten_items(section: dict) -> list[dict]:
    out: list[dict] = []
    for sub in section.get("subsections", []):
        out.extend(sub.get("items", []))
    return out


def _clear_item_row(ws, row: int) -> None:
    for col in INPUT_COLS_TEXT:
        ws[f"{col}{row}"] = ""
    for col in INPUT_COLS_NUM:
        ws[f"{col}{row}"] = None


def _write_item(ws, row: int, item: dict, force_total: bool = False) -> None:
    ws[f"A{row}"] = item.get("desc", "")
    ws[f"B{row}"] = item.get("brand", "")
    ws[f"C{row}"] = item.get("specs", "")
    ws[f"D{row}"] = item.get("qty", 1)
    ws[f"E{row}"] = item.get("pu", 0)
    ws[f"S{row}"] = item.get("note", "")

    typ = item.get("type", "item")
    if typ == "transport":
        ws[f"G{row}"] = item.get("sc", 0)
        ws[f"I{row}"] = 0
        ws[f"K{row}"] = 0
        ws[f"M{row}"] = item.get("ric", 0)
        ws[f"P{row}"] = item.get("impr", 0)
    else:
        ws[f"G{row}"] = item.get("sc", 0)
        ws[f"I{row}"] = item.get("op_acc", 0)
        ws[f"K{row}"] = item.get("inst", 0)
        ws[f"M{row}"] = item.get("ric", 0)
        ws[f"P{row}"] = item.get("impr", 0)

    if force_total:
        ws[f"R{row}"] = item.get("total", 0)


def _set_titles(ws, data: dict, header_row: int) -> None:
    title_row = header_row - 2
    scenario_row = header_row - 1
    coeff_row = header_row + 1

    cliente = data.get("cliente", "")
    indirizzo = data.get("indirizzo", "")
    data_str = data.get("data", "")

    ws[f"A{title_row}"] = f"  STIMA PRELIMINARE — {cliente}  |  {indirizzo}  | "
    ws[f"S{title_row}"] = data_str
    ws[f"A{scenario_row}"] = f"  {data.get('titolo', '')}"
    m_val = ws[f"M{scenario_row}"].value
    s_val = ws[f"S{scenario_row}"].value
    a_val = ws[f"A{coeff_row}"].value
    if m_val is not None and str(m_val).strip() != "":
        ws[f"M{scenario_row}"] = data.get("note_header", "")
    if s_val is not None and str(s_val).strip() != "":
        ws[f"S{scenario_row}"] = data.get("summary_header", "")
    if a_val is not None and str(a_val).strip() != "":
        ws[f"A{coeff_row}"] = "  Coefficienti individuali per riga (non uniformi) — vedi colonne I, K, M, P"


def _set_section_headers(
    ws,
    sections: list[dict],
    section_rows: list[int],
    subsection_rows_by_section: list[list[int]],
) -> None:
    for sec_idx, sec in enumerate(sections):
        existing = str(ws[f"A{section_rows[sec_idx]}"].value or "")
        if "━━━━━━" in existing:
            ws[f"A{section_rows[sec_idx]}"] = f"  ━━━━━━  {sec.get('label', '')}  ━━━━━━"
        elif "------" in existing:
            ws[f"A{section_rows[sec_idx]}"] = f"  ------  {sec.get('label', '')}  ------"
        else:
            ws[f"A{section_rows[sec_idx]}"] = f"  {sec.get('label', '')}"
        for row, sub in zip(subsection_rows_by_section[sec_idx], sec.get("subsections", [])):
            ws[f"A{row}"] = f"  {sub.get('label', '')}"


def _populate_sections(
    ws,
    sections: list[dict],
    section_rows: list[int],
    section_end_rows: list[int],
) -> None:
    for sec_idx, section in enumerate(sections):
        items = _flatten_items(section)
        prefer_lumpsum = any((it.get("type") == "lumpsum") for it in items)
        rows = _find_item_rows_in_section(ws, section_rows[sec_idx], section_end_rows[sec_idx], prefer_lumpsum)

        for row in rows:
            _clear_item_row(ws, row)

        for row, item in zip(rows, items):
            _write_item(ws, row, item, force_total=(item.get("type") == "lumpsum"))


def _is_external_candidate_row(ws, row: int) -> bool:
    a_val = str(ws[f"A{row}"].value or "")
    a_norm = _norm(a_val)
    if any(x in a_norm for x in ("sezione", "subtotale", "totale", "esclusioni")):
        return False
    if a_val.lstrip().startswith("  "):
        return False
    d_val = ws[f"D{row}"].value
    e_val = ws[f"E{row}"].value
    f_val = ws[f"F{row}"].value
    return (d_val is not None) or (e_val is not None) or (f_val is not None)


def _find_external_rows(ws, search_start: int) -> list[int]:
    rows: list[int] = []

    ex_row = _find_row_by_a_contains(ws, "ESCLUSIONI", start=search_start)
    start = ex_row + 1 if ex_row else search_start

    for idx in range(start, ws.max_row + 1):
        if _is_external_candidate_row(ws, idx):
            rows.append(idx)

    # backward-compat: template vecchio con una sola riga external fuori blocco "ESCLUSIONI"
    if not rows:
        first = _find_row_by_a_contains(ws, "Connessione E-Distribuzione", start=search_start)
        if first:
            rows = [first]

    return rows


def _populate_external(ws, external: Iterable[dict], search_start: int) -> dict:
    ext_list = list(external)
    rows = _find_external_rows(ws, search_start)
    if not rows:
        return {"count": len(ext_list), "supported_rows": 0, "overflow": len(ext_list)}

    for row in rows:
        _clear_item_row(ws, row)

    for row, item in zip(rows, ext_list):
        _write_item(ws, row, item, force_total=False)

    overflow = max(0, len(ext_list) - len(rows))
    return {"count": len(ext_list), "supported_rows": len(rows), "overflow": overflow}


def _populate_riepilogo(ws, data: dict) -> None:
    cliente = data.get("cliente", "")
    data_str = data.get("data", "")
    ws["A1"] = f"  RIEPILOGO — {cliente} — Stima Preliminare {data_str}"

    econ = data.get("economic", {})
    ct30 = data.get("ct30", {})

    if "risparmio_annuo" in econ:
        ws["D22"] = econ["risparmio_annuo"]
    if "payback" in econ:
        ws["D23"] = econ["payback"]
    if "van_20a" in econ:
        ws["D24"] = econ["van_20a"]

    if "tecnologia" in ct30:
        ws["B28"] = ct30["tecnologia"]
    if "ci" in ct30:
        ws["B29"] = f"{ct30['ci']} €/kWh"
    if "prated" in ct30:
        ws["B30"] = f"{ct30['prated']} kW EN 14825"
    if "quf" in ct30:
        ws["B31"] = f"{ct30['quf']} h/anno equivalenti (stima)"
    if "kp" in ct30:
        ws["B32"] = str(ct30["kp"])
    if "ei" in ct30:
        ws["B33"] = f"{ct30['ei']} kWh/anno"
    if "itot" in ct30:
        ws["B35"] = f"~{ct30['itot']} € STIMA ±30% — calcolo formale GSE obbligatorio"
        ws["B36"] = f"~{round(ct30['itot'] / 5, 2)} €/anno"


def inspect_template_capacity(data: dict, template_path: Path | None = None) -> dict:
    template = template_path or DEFAULT_TEMPLATE
    wb = load_workbook(template)
    ws = wb[SCENARIO_SHEET]

    sections = data.get("sections", [])
    header_row = _find_header_row(ws)
    section_rows = _find_section_rows(ws, sections) if sections else []

    section_info = []
    for idx, sec in enumerate(sections):
        sec_row = section_rows[idx]
        next_sec = section_rows[idx + 1] if idx + 1 < len(section_rows) else None
        sec_end = _find_section_subtotal_row(ws, sec_row, next_section_row=next_sec)
        items = _flatten_items(sec)
        prefer_lumpsum = any((it.get("type") == "lumpsum") for it in items)
        item_rows = _find_item_rows_in_section(ws, sec_row, sec_end, prefer_lumpsum)
        section_info.append(
            {
                "index": idx,
                "label": sec.get("label", ""),
                "item_count": len(items),
                "capacity": len(item_rows),
                "overflow": max(0, len(items) - len(item_rows)),
                "section_row": sec_row,
                "subtotal_row": sec_end,
            }
        )

    external_start = section_info[-1]["subtotal_row"] + 1 if section_info else (header_row + 1)
    external_rows = _find_external_rows(ws, external_start)
    external_items = len(data.get("external", []))

    return {
        "template": str(template),
        "scenario_sheet": SCENARIO_SHEET,
        "header_row": header_row,
        "sections": section_info,
        "external": {
            "count": external_items,
            "supported_rows": len(external_rows),
            "overflow": max(0, external_items - len(external_rows)),
            "rows": external_rows,
        },
    }


def build(
    data: dict,
    template_path: Path | None = None,
    expected_template_version: str | None = None,
):
    template = template_path or DEFAULT_TEMPLATE
    if not template.exists():
        raise FileNotFoundError(f"Template non trovato: {template}")

    wb = load_workbook(template)
    template_version = _get_defined_name_value(wb, TEMPLATE_VERSION_NAME)
    if expected_template_version and template_version != expected_template_version:
        raise ValueError(
            f"Template version mismatch: atteso '{expected_template_version}', trovato '{template_version}'"
        )
    if template_version:
        _set_defined_name_value(wb, TEMPLATE_VERSION_NAME, template_version)
    _set_defined_name_value(wb, OUTPUT_GENERATOR_NAME, GENERATOR_VERSION)

    ws = wb[SCENARIO_SHEET]
    wr = wb[RIEPILOGO_SHEET]

    sections = data.get("sections", [])
    header_row = _find_header_row(ws)
    _set_titles(ws, data, header_row)

    if sections:
        section_rows = _find_section_rows(ws, sections)
        section_end_rows: list[int] = []
        subsection_rows_by_section: list[list[int]] = []

        for idx, sec_row in enumerate(section_rows):
            next_sec = section_rows[idx + 1] if idx + 1 < len(section_rows) else None
            sec_end = _find_section_subtotal_row(ws, sec_row, next_section_row=next_sec)
            section_end_rows.append(sec_end)
            subsection_rows_by_section.append(
                _find_subsection_rows(ws, sec_row, sec_end, sections[idx].get("subsections", []))
            )

        _set_section_headers(ws, sections, section_rows, subsection_rows_by_section)
        _populate_sections(ws, sections, section_rows, section_end_rows)

        external_start = section_end_rows[-1] + 1
    else:
        external_start = header_row + 1

    _populate_external(ws, data.get("external", []), external_start)
    _populate_riepilogo(wr, data)
    return wb


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("input_json", nargs="?", help="Path JSON input")
    parser.add_argument("output_xlsx", nargs="?", default="COMPUTO_OUTPUT.xlsx", help="Path XLSX output")
    parser.add_argument("template_xlsx", nargs="?", default=str(DEFAULT_TEMPLATE), help="Path template XLSX")
    parser.add_argument(
        "--template",
        dest="template_override",
        help="Override template path (used with --stamp-template-version)",
    )
    parser.add_argument(
        "--template-version",
        dest="expected_template_version",
        help="Fail if template version differs",
    )
    parser.add_argument(
        "--stamp-template-version",
        dest="stamp_template_version_value",
        help=f"Write/update {TEMPLATE_VERSION_NAME} inside template and exit",
    )

    args = parser.parse_args()
    template_xlsx = Path(args.template_override or args.template_xlsx)

    if args.stamp_template_version_value:
        stamp_template_version(template_xlsx, args.stamp_template_version_value)
        print(f"Stamped {TEMPLATE_VERSION_NAME}={args.stamp_template_version_value} in {template_xlsx}")
        if not args.input_json:
            sys.exit(0)

    if not args.input_json:
        parser.error("input_json is required unless using only --stamp-template-version")

    input_json = Path(args.input_json)
    output_xlsx = Path(args.output_xlsx)

    with input_json.open(encoding="utf-8") as f:
        payload = json.load(f)

    workbook = build(payload, template_xlsx, expected_template_version=args.expected_template_version)
    workbook.save(output_xlsx)
    print(f"Saved: {output_xlsx}")
